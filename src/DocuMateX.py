"""
====================================================================
 DocuMateX — Batched Word Mail Merge Edition
====================================================================

 DocuMateY proved the concept: Python orchestrating Word's
 native Mail Merge engine instead of rendering documents itself.

 But Y had a problem. When processing large volumes (1000+
 records), Word's single Execute() call would hang. It was
 trying to build a massive document entirely in memory,
 and the COM interface became unresponsive.

 X solves this by introducing batched execution.

--------------------------------------------------------------------
 WHAT CHANGED FROM Y TO X
--------------------------------------------------------------------

 Y: Single Execute() for the entire record range.
    Fast for small batches. Hangs on large volumes.

 X: Splits the range into batches of 250 records.
    Each batch produces a temp .docx file.
    After all batches complete, they are combined
    into one final document.

 Everything else remains identical:
   - load_data()           → same
   - filter_records()      → same (5-step validation gauntlet)
   - format_dates()        → same
   - update_excel_status() → same
   - run()                 → same flow, same prompts

--------------------------------------------------------------------
 HOW BATCHING WORKS
--------------------------------------------------------------------

 1. Python calculates the full record range (e.g. 1 to 1440).

 2. The range is split into batches of 250:
    Batch 1: records 1-250
    Batch 2: records 251-500
    Batch 3: records 501-750
    ...and so on.

 3. For each batch, Word's Mail Merge executes with that
    specific From/To range. The result is saved as a
    temporary .docx file.

 4. Once all batches are done, the temp files are combined
    into one final document using Word's InsertFile with
    section breaks between batches.

 5. Temp files are cleaned up automatically.

--------------------------------------------------------------------
 PERFORMANCE
--------------------------------------------------------------------

 1,440 records:
   - Manual process:     ~126 hours
   - DocuMate v3 (docxtpl): ~3.5 minutes
   - DocuMateY (single MM):  hangs on this volume
   - DocuMateX (batched MM): ~52 seconds

 The batching approach turned a hanging system into the
 fastest version of DocuMate yet.

--------------------------------------------------------------------
 KNOWN LIMITATIONS
--------------------------------------------------------------------

 - Filtered IN PROCESS records must be continuous in the
   Excel sheet. Scattered rows would include unwanted
   records in between.
 - Word's "Select Table" dialog may appear on template open
   if the template has a saved data source link.
 - Requires Microsoft Word desktop installed.
 - Windows only (win32com).
 - Temp files are written to output_files/ during processing
   and cleaned up after. If the process crashes mid-batch,
   leftover temp files may remain.

--------------------------------------------------------------------
 REQUIREMENTS
--------------------------------------------------------------------

 pip install pywin32 pandas openpyxl
 Microsoft Word desktop must be installed.

====================================================================
"""

import pandas as pd
import time
import os
import sys
import ctypes
from openpyxl import load_workbook
from datetime import datetime

try:
    import pythoncom
    import win32com.client
except ImportError:
    print("❌ pywin32 is required for Mail Merge mode.")
    print("   Install it with: pip install pywin32")
    sys.exit(1)


# Word constants
WD_SEND_TO_NEW_DOCUMENT = 0
WD_DO_NOT_SAVE_CHANGES = 0
WD_FORMAT_DOCUMENT_DEFAULT = 16   # .docx


class BirthRegistrationProcessor:
    """
    Same structure as original DocuMate v3.
    Only generate_and_merge_documents() is changed to Word Mail Merge.
    """

    def __init__(self, excel_path: str, sheet_name: str, template_path: str,
                 output_folder: str, update_existing: bool = False):
        self.excel_path = excel_path
        self.sheet_name = sheet_name
        self.template_path = template_path
        self.output_folder = output_folder
        self.data = None
        self.update_existing = update_existing

    def load_data(self) -> None:
        """
        Step 1: Read Excel into pandas for validation and filtering.
        """
        self.data = pd.read_excel(self.excel_path, sheet_name=self.sheet_name)

    def filter_records(self) -> None:
        """
        Step 2: Filter and validate records.

        This is the big upgrade from v2. Instead of just filtering
        by status, v3 runs a gauntlet of checks before allowing
        processing to continue:

        Check 1 → Is there any data at all?
        Check 2 → Does the 'Status' column exist?
        Check 3 → Are there any "IN PROCESS" records?
        Check 4 → Do all mandatory fields have values?
        Check 5 → Are there any duplicate File_Numbers?

        If ANY check fails, processing stops with a clear message.
        self.data is set to None so downstream methods know to skip.
        """

        # --- Check 1: Is there any data at all? ---
        if self.data is None or self.data.empty:
            msg = (
                "📝 Found no records to populate.\n\n"
                "🧠 Powered by ENN — Fuelled by curiosity, refined by data-driven clarity."
            )
            ctypes.windll.user32.MessageBoxW(0, msg, "DocuMate", 0)
            self.data = None
            return

        # --- Check 2: Does the 'Status' column exist? ---
        if "STATUS" not in self.data.columns:
            msg = (
                "⚠️ 'Status' column not found. Please check the Excel file.\n\n"
                "🧠 Powered by ENN — Fuelled by curiosity, refined by data-driven clarity."
            )
            ctypes.windll.user32.MessageBoxW(0, msg, "DocuMate", 0)
            self.data = None
            return

        # --- Check 3: Filter to "IN PROCESS" only ---
        self.data = self.data[self.data["STATUS"].str.upper() == "IN PROCESS"]

        if self.data.empty:
            msg = (
                "📝 No records found with status 'In Process'.\n\n"
                "🧠 Powered by ENN — Fuelled by curiosity, refined by data-driven clarity."
            )
            ctypes.windll.user32.MessageBoxW(0, msg, "DocuMate", 0)
            self.data = None
            return

        # --- Check 4: Are any mandatory fields blank or missing? ---
        # Skips columns that are allowed to be empty
        # (Status, Date_Printed, Date_Issued are not mandatory)
        cols_to_check = [col for col in self.data.columns if col not in ["STATUS", "Date_Printed", "Date_Issued"]]
        missing_mask = (
            self.data[cols_to_check].isna()
            | self.data[cols_to_check].apply(lambda col: col.astype(str).str.strip() == "")
        )

        rows_with_missing = self.data[missing_mask.any(axis=1)]

        if not rows_with_missing.empty:
            # Print which rows have problems and which fields are missing
            print("\nDocuMate :⚠️ Records with missing mandatory fields detected:\n")

            for idx, row in rows_with_missing.iterrows():
                missing_cols = [
                    col for col in cols_to_check
                    if pd.isna(row[col]) or str(row[col]).strip() == ""
                ]
                # +2 because: row 1 = header, row index starts at 0
                print(f"➡️ Row {idx + 2} | Missing fields: {', '.join(missing_cols)}")

            msg = (
                f"⚠️ Warning: Found {len(rows_with_missing)} record(s) with missing values.\n"
                "Please correct the highlighted rows in the Excel file before proceeding.\n\n"
                "🧠 Powered by ENN — Fuelled by curiosity, refined by data-driven clarity."
            )
            ctypes.windll.user32.MessageBoxW(0, msg, "DocuMate", 0)
            self.data = None
            return

        # --- Check 5: Are there duplicate File_Numbers? ---
        # Duplicates could mean the same person was entered twice,
        # which would produce duplicate certificates
        duplicates = self.data[self.data.duplicated(subset=["File_Number"], keep=False)]
        if not duplicates.empty:
            print(duplicates[["File_Number","Serial","Name","Date_Issued"]])
            msg = (
                f"⚠️ Warning: Found {len(duplicates)} duplicate File_Number entries in 'In Process' records:\n"
                "Please check the Excel file for duplicates.\n\n"
                "🧠 Powered by ENN — Fuelled by curiosity, refined by data-driven clarity."
            )
            ctypes.windll.user32.MessageBoxW(0, msg, "DocuMate", 0)
            self.data = None
            return

    def format_dates(self, date_columns):
        """
        Step 3: Clean up date columns.
        Same as v2 — converts to dd/mm/yyyy for clean document output.
        """
        for col in date_columns:
            if col in self.data.columns:
                self.data[col] = pd.to_datetime(self.data[col], errors="coerce").dt.strftime("%d/%m/%Y")

    def generate_and_merge_documents(self, merged_filename=f"DocuMateX_BIRTH_REGISTRATION_{datetime.now().strftime('%d%m%Y')}.docx") -> None:
        """
        Step 4: Generate and merge all documents using Microsoft Word Mail Merge
        directly against the original Excel datasource.

        Changes from GPT's original:
        - DispatchEx instead of Dispatch (forces fresh Word instance)
        - AutomationSecurity = 3 (suppresses data source confirmation dialogs)
        - Batched execution (250 records per batch) to prevent Word hanging on large volumes
        - Debug prints before/after each critical step to identify hang points
        - Temp batch files cleaned up after final merge
        """

        if self.data is None or self.data.empty:
            print("DocuMate : ⚠️ No records found for merging.")
            return

        # Reload full source to map filtered rows back to Word mail merge record positions
        full_df = pd.read_excel(self.excel_path, sheet_name=self.sheet_name)
        full_df["__merge_record__"] = full_df.index + 1

        # Sort by Serial
        if "Serial" in self.data.columns:
            try:
                self.data["__serial_num__"] = (
                    self.data["Serial"].astype(str).str.extract(r"(\d+)", expand=False).fillna("0").astype(int)
                )
                self.data = self.data.sort_values("__serial_num__")
            except Exception:
                print("⚠️ Serial sort skipped.")

        # Map File_Number to original Excel row position
        merge_map = dict(
            zip(full_df["File_Number"].astype(str).str.strip(), full_df["__merge_record__"])
        )

        self.data["__merge_record__"] = (
            self.data["File_Number"].astype(str).str.strip().map(merge_map)
        )

        if self.data["__merge_record__"].isna().any():
            raise ValueError("Some filtered records could not be mapped back to original Excel rows.")

        merge_records = sorted(self.data["__merge_record__"].astype(int).tolist())

        first_record = min(merge_records)
        last_record = max(merge_records)

        # Direct From/To is safe only when records are continuous
        expected = list(range(first_record, last_record + 1))
        if merge_records != expected:
            raise ValueError(
                "Filtered IN PROCESS records are not continuous in the Excel source. "
                "Direct Word Mail Merge range would include unwanted rows in between."
            )

        total_records = len(self.data)
        print(f"DocuMate : ⏳ Processing {total_records} records with Word Mail Merge...")
        print(f"DocuMate : 📌 Auto-selected range in Word: From {self.data.iloc[0]['__serial_num__']} To {self.data.iloc[-1]['__serial_num__']}")
        if not os.path.exists(self.output_folder):
            print(f"DocuMate : 📂 Creating output folder '{self.output_folder}'...")
            os.makedirs(self.output_folder)

        excel_abs = os.path.abspath(self.excel_path)
        template_abs = os.path.abspath(self.template_path)
        output_path = os.path.abspath(os.path.join(self.output_folder, merged_filename))

        if not os.path.exists(excel_abs):
            raise FileNotFoundError(f"Excel source not found: {excel_abs}")
        if not os.path.exists(template_abs):
            raise FileNotFoundError(f"Template not found: {template_abs}")

        # Batch setup - Word chokes on 1000+ records in a single Execute
        batch_size = 250
        batches = []
        for i in range(first_record, last_record + 1, batch_size):
            batch_end = min(i + batch_size - 1, last_record)
            batches.append((i, batch_end))

        word = None
        main_doc = None
        final_doc = None
        temp_files = []

        try:
            pythoncom.CoInitialize()

            print("DocuMate : 🔧 Starting Word Mail Merge...")
            word = win32com.client.DispatchEx("Word.Application")
            word.Visible = True
            word.DisplayAlerts = 0
            word.AutomationSecurity = 3  # msoAutomationSecurityForceDisable

            time.sleep(1)

            main_doc = word.Documents.Open(template_abs, ReadOnly=True, AddToRecentFiles=False)
            time.sleep(0.5)

            sheet = self.sheet_name.replace("'", "''")

            print("DocuMate : 🔗 Connecting data source,Please approve the action in Word file if prompted and wait for the document generation to complete...")
            main_doc.MailMerge.OpenDataSource(
                Name=excel_abs,
                ConfirmConversions=False,
                ReadOnly=True,
                LinkToSource=True,
                AddToRecentFiles=False,
                Revert=False,
                Connection=(
                    "Provider=Microsoft.ACE.OLEDB.12.0;"
                    f"Data Source={excel_abs};"
                    'Extended Properties="Excel 12.0 Xml;HDR=YES;IMEX=1";'
                ),
                SQLStatement=f"SELECT * FROM [{sheet}$]",
                SubType=0
            )
            print("DocuMate : ✅ Data source connected successfully, Please wait while I generate the documents...")

            time.sleep(0.5)

            main_doc.MailMerge.MainDocumentType = 0
            main_doc.MailMerge.Destination = WD_SEND_TO_NEW_DOCUMENT
            main_doc.MailMerge.SuppressBlankLines = True

            merge_start = time.time()

            # Process each batch
            for batch_num, (batch_start, batch_end) in enumerate(batches, start=1):
                main_doc.MailMerge.DataSource.FirstRecord = int(batch_start)
                main_doc.MailMerge.DataSource.LastRecord = int(batch_end)
                main_doc.MailMerge.Execute(Pause=False)

                time.sleep(0.5)

                merged_doc = word.ActiveDocument

                temp_path = os.path.abspath(
                    os.path.join(self.output_folder, f"__temp_batch_{batch_num}.docx")
                )
                merged_doc.SaveAs2(temp_path, FileFormat=WD_FORMAT_DOCUMENT_DEFAULT)
                temp_files.append(temp_path)

                try:
                    merged_doc.Close(False)
                except Exception:
                    pass
                merged_doc = None


            # Close the template before combining
            try:
                main_doc.Close(False)
            except Exception:
                pass
            main_doc = None

            # Combine all batch files into one final document
            if len(temp_files) == 1:
                # Only one batch, just rename it
                import shutil
                shutil.move(temp_files[0], output_path)
                temp_files = []
            else:

                final_doc = word.Documents.Open(os.path.abspath(temp_files[0]))

                for i, temp_file in enumerate(temp_files[1:], start=2):
                    rng = final_doc.Range(final_doc.Content.End - 1, final_doc.Content.End - 1)
                    rng.InsertBreak(7)  # wdSectionBreakNextPage
                    rng = final_doc.Range(final_doc.Content.End - 1, final_doc.Content.End - 1)
                    rng.InsertFile(os.path.abspath(temp_file))

                final_doc.SaveAs2(output_path, FileFormat=WD_FORMAT_DOCUMENT_DEFAULT)

            print(f"\nDocuMate : ✅ Generated {total_records} documents saved in:\n📂 {output_path}")

        except Exception as e:
            print(f"\n⚠️ DocuMate : Mail Merge error — {e}")
            raise

        finally:
            # Close everything safely
            try:
                if final_doc is not None:
                    final_doc.Close(False)
            except Exception:
                pass

            try:
                if main_doc is not None:
                    main_doc.Close(False)
            except Exception:
                pass

            try:
                if word is not None:
                    word.Quit()
            except Exception:
                pass

            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

            # Clean temp files
            for f in temp_files:
                try:
                    if os.path.exists(f):
                        os.remove(f)
                except Exception:
                    pass

            # Clean helper columns
            if self.data is not None:
                for col in ["__serial_num__", "__merge_record__"]:
                    if col in self.data.columns:
                        self.data.drop(columns=[col], inplace=True, errors="ignore")

    def update_excel_status(self) -> None:
        """
        Step 5 (optional): Mark processed records as PRINTED in Excel.
        Same original logic: all current IN PROCESS rows get updated.
        """
        try:
            wb = load_workbook(self.excel_path)
            ws = wb[self.sheet_name]

            headers = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}
            if "STATUS" not in headers:
                print("⚠️ STATUS column not found — skipping Excel update.")
                return

            if "Date_Printed" not in headers:
                next_col = ws.max_column + 1
                ws.cell(row=1, column=next_col, value="Date_Printed")
                headers["Date_Printed"] = next_col

            today = datetime.now().strftime("%d/%m/%Y")

            status_col = headers["STATUS"]
            date_col = headers["Date_Printed"]
            in_process_rows = [
                r for r in range(2, ws.max_row + 1)
                if str(ws.cell(row=r, column=status_col).value).strip().upper() == "IN PROCESS"
            ]

            for r in in_process_rows:
                ws.cell(row=r, column=status_col, value="PRINTED")
                ws.cell(row=r, column=date_col, value=today)

            wb.save(self.excel_path)
            print(f"🗂️ DocuMate : Sheet '{self.sheet_name}' updated successfully on {today} at {time.strftime('%H:%M:%S')}.")

        except PermissionError:
            print("❌ Cannot update Excel while file is open. Please close it and retry.")
        except Exception as e:
            print(f"⚠️ DocuMate : Error updating Excel — {e}")

    def run(self) -> None:
        """
        Same original flow.
        """
        print("Deploying DocuMate...\n")
        print("DocuMate : Alright ENN, Ready to engage - initializing populate launch 🚀\n")
        try:
            print("DocuMate : 📄 Reading Table...")
            self.load_data()

            print("DocuMate : 🔍 Searching for new applicants...")
            self.filter_records()
            if self.data is None:
                return

            # --- Format dates ---
            self.format_dates(["Registration_date"])
            print(f"DocuMate : 📝 Found {len(self.data)} new creations of Lord Bramha 🙏🏻...")

            total_start = time.time()

            print("DocuMate : ⚙️  Initializing Microsoft Word Mail Merge engine...")
            merge_start = time.time()
            self.generate_and_merge_documents()
            merge_time = time.time() - merge_start
            print(f"\nDocuMate :✅ Generated and merged {len(self.data)} documents")
            print(f"DocuMate :⚡ Word Mail Merge completed in {merge_time:.2f} seconds.")

            user_choice = input(
                f"DocuMate : Do you want me to mark the statuses as PRINTED for {len(self.data)} records? (yes/no): "
            ).strip().lower()
            self.update_existing = user_choice in ("yes", "y")

            update_text = ""
            excel_time = 0.0
            if self.update_existing:
                print("\nDocuMate : ⏳ Updating Excel statuses...\n")
                excel_start = time.time()
                self.update_excel_status()
                excel_time = time.time() - excel_start
                update_text = f" and updated Excel in {excel_time:.2f} seconds"

            total_time = time.time() - total_start

            msg = (
                f"✅ DocuMate : Mission accomplished!\n"
                f"✅ Generated {len(self.data)} documents successfully{update_text}.\n"
                f"✅ Total time taken: {total_time:.2f} seconds.\n\n"
                "⚡ Powered by Word Mail Merge + Python.\n"
                "🧠 Powered by ENN — Fuelled by curiosity, refined by data-driven clarity."
            )
            ctypes.windll.user32.MessageBoxW(0, msg, "DocuMate", 0)

        except PermissionError:
            msg = (
                "❌ I cannot engage the populate launch because the Excel file is open.\n"
                "Please close it and try again later.\n\nMission aborted."
            )
            ctypes.windll.user32.MessageBoxW(0, msg, "DocuMate", 0)

        except Exception as e:
            msg = (
                f"⚠️ DocuMate encountered an unexpected error:\n{e}\n\n"
                "🧠 Powered by ENN — Fuelled by curiosity, refined by data-driven clarity."
            )
            ctypes.windll.user32.MessageBoxW(0, msg, "DocuMate", 0)


if __name__ == "__main__":
    base_dir = (
        os.path.dirname(os.path.dirname(sys.executable))
        if getattr(sys, "frozen", False)
        else os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    )

    docuMatePLUSAgent = BirthRegistrationProcessor(
        excel_path=os.path.join(base_dir, "data", "DocuMate_DataFrame.xlsx"),
        sheet_name="DocuMateSRC",
        template_path=os.path.join(base_dir, "templates", "DOCUMENT_TEMPLATE_FILE_MM.docx"),
        output_folder=os.path.join(base_dir, "output_files"),
        update_existing=False,
    )

    docuMatePLUSAgent.run()